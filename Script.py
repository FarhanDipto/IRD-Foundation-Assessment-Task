# Assalamualaikum!
# Please use this command to run:: python Script.py OCR.docx OCR_parsed.xlsx
# Thank you!

import re
import sys
from pathlib import Path
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# --------- Mappings ----------
BENGALI_DIGITS_MAP = {ord(x): str(i) for i, x in enumerate("০১২৩৪৫৬৭৮৯")}
def bengali_to_int(bengali_str):
    if bengali_str is None:
        return None
    ascii_digits = bengali_str.translate(BENGALI_DIGITS_MAP)
    ascii_digits = re.sub(r'[^\d]', '', ascii_digits)
    return int(ascii_digits) if ascii_digits else None

# CHAPTER_RE = re.compile(r'^\s*অধ্যা[যয়]?\s*:?\s*(?P<name>.+)$')  
CHAPTER_RE = re.compile(r'^\s*অধ্যায়\s*:?\s*(?P<name>.+)$')
HADITH_RE  = re.compile(r'^\s*\[(?P<num>[\u09E6-\u09EF0-9]+)\]\s*(?P<body>.*)$')

SENTENCE_END_CHARS = ('।', '.', '?', '!', '।"', '?"', '!"')

# --------- Parsing logic ----------
def parse_docx(path):
    doc = Document(path)
    paras = [p.text for p in doc.paragraphs]
    n = len(paras)

    chapter_name = None
    sections = []
    hadiths = []
    current_hadith = None

    def finalize_hadith():
        nonlocal current_hadith, hadiths
        if current_hadith:
            text = "\n\n".join([p for p in current_hadith['paras'] if p != ""])
            hadiths.append({'id': current_hadith['id'], 'hadith': text, 'raw_id': current_hadith.get('raw_id')})
            current_hadith = None

    def next_nonempty(i):
        j = i+1
        while j < n and paras[j].strip() == "":
            j += 1
        if j < n:
            return j, paras[j].strip()
        return None, None

    for i, raw in enumerate(paras):
        text = raw.strip()
        if text == "":
            if current_hadith:
                current_hadith['paras'].append("")  
            continue

        # Chapter?
        m = CHAPTER_RE.match(text)
        if m:
            finalize_hadith()
            chapter_name = m.group('name').strip()
            continue

        # Hadith start?
        m2 = HADITH_RE.match(text)
        if m2:
            finalize_hadith()
            raw_num = m2.group('num')
            num = bengali_to_int(raw_num)
            body = m2.group('body').strip()
            current_hadith = {'id': num, 'raw_id': raw_num, 'paras': []}
            if body:
                current_hadith['paras'].append(body)
            continue

        # Section (heuristic)
        is_short = len(text) <= 120
        ends_with_sentence = any(text.endswith(ch) for ch in SENTENCE_END_CHARS)
        contains_brackets = '[' in text and ']' in text
        nxt_idx, nxt_text = next_nonempty(i)
        next_is_hadith = False
        if nxt_text and HADITH_RE.match(nxt_text):
            next_is_hadith = True

        treat_section = False
        if is_short and not ends_with_sentence and not contains_brackets:
            if current_hadith is None:
                treat_section = True
            else:
                # Detecting section inside a hadith
                if next_is_hadith:
                    treat_section = True

        if treat_section:
            finalize_hadith()
            sec = text
            if not sections or sections[-1] != sec:
                sections.append(sec)
            continue

        if current_hadith is not None:
            current_hadith['paras'].append(text)
        else:
            pass

    finalize_hadith()
    return chapter_name, sections, hadiths

# --------- Excel writing ----------
def write_xlsx(path, chapter_name, sections, hadiths):
    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    # chapter
    sch = wb.create_sheet("chapter")
    sch.append(["id", "name"])
    sch["A1"].font = Font(bold=True)
    sch["B1"].font = Font(bold=True)
    sch["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sch["B1"].alignment = Alignment(horizontal="center", vertical="center")
    sch.append([1, chapter_name or ""])
    sch.column_dimensions['A'].width = 8
    sch.column_dimensions['B'].width = 60
    for row in sch.iter_rows(min_row=2, max_col=2):
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[1].alignment = Alignment(wrapText=True, vertical="top")

    # section
    ssec = wb.create_sheet("section")
    ssec.append(["id", "name"])
    ssec["A1"].font = Font(bold=True)
    ssec["B1"].font = Font(bold=True)
    ssec["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ssec["B1"].alignment = Alignment(horizontal="center", vertical="center")
    for idx, s in enumerate(sections, start=1):
        ssec.append([idx, s])
    ssec.column_dimensions['A'].width = 8
    ssec.column_dimensions['B'].width = 60
    for row in ssec.iter_rows(min_row=2, max_col=2):
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[1].alignment = Alignment(wrapText=True, vertical="top")

    # hadith
    sh = wb.create_sheet("hadith")
    sh.append(["id", "hadith"])
    sh["A1"].font = Font(bold=True)
    sh["B1"].font = Font(bold=True)
    sh["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sh["B1"].alignment = Alignment(horizontal="center", vertical="center")
    for h in hadiths:
        sh.append([h['id'], h['hadith']])
    sh.column_dimensions['A'].width = 8
    sh.column_dimensions['B'].width = 80
    for row in sh.iter_rows(min_row=2, max_col=2):
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[1].alignment = Alignment(wrapText=True, vertical="top")

    wb.save(path)

# --------- Main CLI ----------
def main():
    if len(sys.argv) < 3:
        print("Usage: python Script.py OCR.docx OCR_parsed.xlsx")
        sys.exit(1)
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    if not input_path.exists():
        print("Input file not found:", input_path)
        sys.exit(2)
    chapter_name, sections, hadiths = parse_docx(input_path)
    write_xlsx(output_path, chapter_name, sections, hadiths)
    print("Wrote:", output_path)
    print("chapter:", chapter_name)
    print("sections:", len(sections), "hadiths:", len(hadiths))

if __name__ == "__main__":
    main()
